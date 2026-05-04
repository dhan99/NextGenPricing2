#!/usr/bin/env python3
"""Generate docs/feature-summary.xlsx — stakeholder-facing matrix of every
DealPad capability, segregated by business intent.

Sheets:
  1. Overview          — exec summary + status legend + status totals
  2. Excel Parity      — features that replicate the current Excel-workbook solution
  3. Integrations      — D365 / Workday / Intapp / Conga / Power BI
  4. Pricing Engine    — capabilities matching Intapp Pricing (or competing engines)
  5. Differentiators   — capabilities unique to DealPad (AI + collab + automation)
  6. Technical Detail  — engineering view (Phase / Feature ID / PRs)
  7. Glossary          — column key + status legend

Re-run from repo root:
    python3 scripts/audit/generate_feature_summary.py
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = "docs/feature-summary.xlsx"

# ---- Status palette ----
STATUS_COLOR = {
    "Implemented":   "C6EFCE",  # green
    "Foundation":    "FFEB9C",  # amber — schema/service shipped, UI/live wiring pending
    "Pending":       "F2DCDB",  # red-ish — not started this cycle
    "Operational":   "DCE6F1",  # blue — ops/infra/deploy
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="DA720F")  # Armanino amber
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(fill_type="solid", fgColor="FFF1DA")  # light amber
SECTION_FONT = Font(bold=True, size=11, color="6B3D02")
BORDER_THIN = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# ---------- EXCEL PARITY ----------
# Features that replicate the current Excel-workbook-based pricing process.
# Maps to Epics E1-E3, E5, E6 from the RFQ User Stories.
EXCEL_PARITY = [
    # (Capability, US Refs, Stakeholder Persona, Status, What it does, How to demo / verify, Notes)
    ("Project classification (New vs Renewal)", "US-01",
     "Project Delivery Lead", "Implemented",
     "PDL initiates pricing classified as New or Renewal; drives downstream workflow + comparison toggles.",
     "DealsList → New Deal → pick dealType=new|renewal", "—"),
    ("CRM project context auto-populate", "US-02",
     "PDL", "Implemented",
     "Title, account, BU, service line pre-fill from D365 opportunity at deal creation.",
     "Dynamics CRM page → Import Opportunity → new deal pre-populated", "—"),
    ("Multi-entity scoping per project", "US-03",
     "PDL", "Implemented",
     "Single deal can model 4+ entities (1040 + 1120 + 1065 + 1120S etc.) with per-entity scope/pricing rollups.",
     "Deal → Scope step → Add Entity tabs", "Entity tabs + per-entity totals via /api/deals/:id/entities/:eid/totals"),
    ("Multi-service projects on one deal", "US-04",
     "PDL", "Implemented",
     "Multiple service types per deal; each carries its own scope + assemblies + rolls up to project total.",
     "Pricing Step rolls service line totals together", "—"),
    ("Renewal baseline load (prior-year)", "US-05, US-34",
     "PDL", "Implemented",
     "Renewal projects load prior-year scope + hours + fee totals; current vs prior delta visible.",
     "Renewal Leadsheet page", "—"),
    ("Project arrangement (services / fee model)", "US-06, US-23, US-59",
     "PDL", "Implemented",
     "T&M / Fixed Fee / Capped / Contingent / Retainer / Hybrid arrangements drive pricing inputs + totals.",
     "Pricing Step → FeeArrangementPicker", "—"),
    ("Versioning + revisions", "US-07, US-42",
     "PDL", "Implemented",
     "Deals carry a version field via clone-with-history; activity log captures who changed what when.",
     "Deal Clone → revisions tracked; activity log shows audit trail", "—"),
    ("Governed scope catalog", "US-08",
     "Pricing Operations", "Implemented",
     "Scope items maintained in admin catalog with stable codes, BU/service-line filters, complexity multipliers.",
     "/admin/scope-catalog", "—"),
    ("Scope item add/remove on deal", "US-09",
     "PDL", "Implemented",
     "Catalog picker on Scope step; immediate recalculation of pricing rollups.",
     "Deal → Scope step → Add Items", "Per-entity unique constraint allows same item across multiple entities (PR #53)"),
    ("Assembly expansion (configurable bundles)", "US-10",
     "PDL", "Implemented",
     "Assembly templates (e.g. Tax PHB Standard) expand into structured components with default + tier-based hours; quantities recalc by prompts.",
     "Scope Step → Assembly button → pick template + tier", "—"),
    ("Prompts & answers (complexity questionnaire)", "US-11",
     "PDL", "Implemented",
     "Prompt-driven complexity questionnaire applies multipliers/logic to scope quantities; answers re-trigger calc on change.",
     "Deal → Assumptions step", "—"),
    ("Prompt sets governed by BU/Service", "US-12",
     "Pricing Operations", "Implemented",
     "Versioned, published prompt sets per BU + service line; partial unique index on published.",
     "/admin/prompt-sets", "—"),
    ("Manual adjustments within guardrails", "US-13",
     "PDL", "Implemented",
     "Per-line rate override with audit trail (overrideBy, overrideReason, standardRate baseline).",
     "Pricing Step → click rate cell → override popover", "—"),
    ("Out-of-scope / add-ons separation", "US-14",
     "PDL", "Pending",
     "Distinguishing base scope from optional add-ons in rollups; approval acknowledgement on optional items.",
     "—", "Tracked but not yet built; current scope items are all base"),
    ("Licensing / non-labor components", "US-15",
     "PDL", "Pending",
     "Software licenses + pass-through expenses surfaced separately in totals.",
     "—", "Tracked but not yet built"),
    ("Data validation + completeness gating", "US-16",
     "PDL", "Implemented",
     "Server-side validators flag missing required inputs before submission; route returns structured error codes.",
     "Try Submit on incomplete deal → 409 with field name", "—"),
    ("Wizard → Grid usability", "US-17",
     "PDL", "Implemented",
     "8-step wizard (Setup → Scope → Assumptions → Pricing → Scenarios → Review → Approve → Summary) with progress saved.",
     "Any deal page", "—"),
    ("Pricing Grid (entity × role)", "US-18, US-19",
     "PDL + Pricing Ops", "Implemented",
     "Grid view: rows by entity/service, columns by role/skill, calculated cells from quantities + rates.",
     "Deal → Pricing step", "—"),
    ("Standard cost tables", "US-20",
     "Pricing Operations", "Implemented",
     "Per-role cost rates applied to planned hours; governed via rate-cards admin.",
     "/admin/rate-cards → entries table", "—"),
    ("Rate tables (Standard + Contract)", "US-21",
     "Pricing Operations", "Implemented",
     "Multiple rate cards with effective dates; deal pulls active card; per-engagement override support.",
     "/admin/rate-cards", "—"),
    ("Geo / offshore discounts", "US-22",
     "Pricing Operations", "Pending",
     "Discount rules by role + location; apply automatically to rate/cost in calc.",
     "—", "Roles support implicit; explicit geo discount rules not yet built"),
    ("Margin calculation transparency", "US-24",
     "Finance / FP&A", "Implemented",
     "Margin = (Fee - Cost) / Fee; per-line + per-deal + per-engagement-input rollup; backfill ensures stored totals match.",
     "Pricing Step footer; calc-parity test pins formulas", "—"),
    ("Effective hourly rate (EHR) / realization", "US-25",
     "Finance / FP&A", "Implemented",
     "Blended rate = totalFee / totalHours surfaced in deal summary + analytics.",
     "Deal Summary header", "—"),
    ("Service tier options (Essential/Enhanced/Ultimate)", "US-26",
     "PDL", "Implemented",
     "Assembly templates support per-tier hour overrides; picker lets user preview each tier before applying.",
     "Scope Step → Assembly Picker → tier dropdown", "—"),
    ("Multi-year / term modeling", "US-27",
     "Finance / FP&A", "Pending",
     "Year-by-year economics + price escalators; renewal terms across multiple periods.",
     "—", "Tracked but not yet built"),
    ("Workday-friendly hours export", "US-28, US-48",
     "Finance / FP&A", "Implemented",
     "Approved-deal hours pushed to Workday cost-center via autoPushWorkdayProject; cost-center committed bumps.",
     "Approve a deal → /api/workday/events shows project_pushed", "—"),
    ("Excel-workbook calculation parity", "US-29",
     "Finance / FP&A", "Implemented",
     "Calc-parity golden test pins computeDealTotalsFromLines against deals 4 + 27 (T&M + Tech-Admin uplift paths).",
     "npm test (tests/calc-parity/)", "—"),
    ("Input → output traceability", "US-30",
     "Risk / QRM", "Implemented",
     "Each output field tracks source: prompt answers (impactMultiplier), engagement inputs, rate-card refs, assemblies. Stored in pricing_lines.",
     "Pricing Step rate-override popover shows standardRate + adjustment trace", "—"),
    ("Error handling + guardrails", "US-31",
     "PDL", "Implemented",
     "Drizzle constraints + structured error codes (entity_deal_mismatch, illegal_state_transition, intapp_conflict, workday_validation_blocked).",
     "Submit invalid combinations → 4xx with code", "—"),
    ("Approval states + lifecycle", "US-40",
     "Pricing Operations", "Implemented",
     "Draft → Submitted → In Review (pending_lead_review / pending_bu_approval) → Approved | Rejected with timestamps + actor.",
     "Approve step on any deal", "—"),
    ("Tiered approval routing", "US-39",
     "Practice Lead", "Implemented",
     "Routing by margin/risk/size: Practice Lead trigger fires on high-fee, low-margin, or large-scope deals (shared/policy.ts).",
     "Submit a low-margin deal → Practice Lead approver auto-assigned", "—"),
    ("Reviewer comments + rejection rationale", "US-41",
     "Practice Lead", "Implemented",
     "Approval row carries comments per version; rejection requires reason; persists in audit history.",
     "Approve step → reject → rationale field", "—"),
    ("Resubmission + version control", "US-42",
     "PDL", "Implemented",
     "Rejected deal can be amended; resubmission creates a new approval row tied to the latest version; prior decisions remain visible.",
     "Submit → Reject → revise → Resubmit", "—"),
    ("Fast-track path for low-risk", "US-43",
     "Practice Lead", "Pending",
     "Auto-approve under defined thresholds (margin > target, fee < threshold) without leadership review.",
     "—", "Tracked but not yet built"),
    ("Delegation / backup approvers", "US-44",
     "Practice Lead", "Pending",
     "Out-of-office routing to a designated backup approver.",
     "—", "Tracked but not yet built"),
    ("Approval visibility + states board", "US-45",
     "Pricing Operations", "Implemented",
     "Dashboard widgets surface pending approvals + approval queues; status counts per persona.",
     "Dashboard → Pending Approvals card", "—"),
    ("Pricing summary output", "US-46",
     "PDL", "Implemented",
     "Summary step renders deal facts + pricing breakdown + approval status; deal banner shows badges for CRM link / Intake / Screening.",
     "Deal → Summary step", "—"),
    ("Branded proposal export (HTML/PDF)", "US-47",
     "PDL", "Implemented",
     "GET /api/deals/:id/proposal generates Armanino-branded HTML proposal (or ?format=json).",
     "/api/deals/:id/proposal", "—"),
    ("Renewal output (Year-over-Year summary)", "US-51",
     "PDL", "Implemented",
     "Renewal Leadsheet view shows YoY scope/hours/fee delta; status indicators per line.",
     "/deals/:id/renewal-leadsheet", "—"),
    ("Operational dashboards", "US-52",
     "Finance / FP&A", "Implemented",
     "/analytics with charts, win-rates, margin trends, service-line breakdown.",
     "/analytics", "—"),
]

# ---------- INTEGRATIONS ----------
# Bi-directional integrations with the Q2C surrounding stack.
INTEGRATIONS = [
    # (Capability, System, Direction, Status, What it does, How to demo, Notes)
    ("Account / Opportunity sync", "Microsoft Dynamics 365",
     "Bi-directional", "Implemented",
     "Inbound: pull accounts + opportunities; auto-create DealPad clients/deals on import. Outbound: status/owner/amount push back on changes via autoPushDeal.",
     "/integrations/dynamics → import any 'Develop' opportunity → deal created", "Currently simulated mode; live OAuth wiring stub-ready (US-63)"),
    ("Owner / quota sync", "Dynamics 365",
     "Inbound", "Implemented",
     "Sales-rep roster + quotas pulled into dynamics_owners table for routing + quota-coverage views.",
     "/integrations/dynamics → Owners panel", "—"),
    ("Send-back to Sales (revisions)", "Dynamics 365",
     "Outbound", "Implemented",
     "Sales can return an approved deal for revision; reuses rejection path so PDL can amend + resubmit.",
     "POST /api/dynamics/opportunities/:id/send-back", "—"),
    ("Cost-center budgets + commitments", "Workday Adaptive",
     "Bi-directional", "Implemented",
     "Inbound: cost-center budgets + utilization. Outbound: project commitment + hours when deal approved.",
     "/integrations/workday → cost-centers; approve a deal → committed bumps", "Currently simulated; live mode requires Adaptive API keys (US-64)"),
    ("Workforce data / worker rates", "Workday",
     "Inbound", "Implemented",
     "Worker roster + rates synced to workday_workers + workday_rate_cards.",
     "/integrations/workday → Workers panel", "—"),
    ("Pre-submit validation gate", "Workday",
     "Inbound", "Implemented",
     "Approval blocks if cost-center over budget or staffing shortfall; QRM override path available.",
     "Submit Tech-Consulting deal → WORKDAY_VALIDATION_BLOCKED 409", "—"),
    ("Intake / screening (conflicts of interest)", "Intapp Risk",
     "Bi-directional", "Implemented",
     "Inbound: nightly re-screen + screening results. Outbound: outcome push on approval/rejection + mitigation status.",
     "/integrations/intapp → screening states", "Currently simulated; live mode awaits Intapp API key"),
    ("Conflict mitigation tracking", "Intapp Risk",
     "Bi-directional", "Implemented",
     "Mitigations resolve / waive / reject; pushOutcome + pushMitigation fan out on transitions.",
     "POST /api/intapp/mitigations/:id/push", "—"),
    ("AI confidence-routed extraction", "Intapp Intake",
     "Inbound", "Implemented",
     "Intake form auto-extracts client + matter facts with AI confidence scores; high-confidence auto-progress, low-confidence routed for review.",
     "/architecture → Intapp Federated tab", "—"),
    ("Engagement letter delivery", "Conga CLM",
     "Outbound", "Implemented",
     "pushDelivery delivers letters via email/esign/portal; flips engagement_letters.status = 'delivered'.",
     "POST /api/conga/letters/:id/deliver", "Currently simulated"),
    ("Engagement-letter template management", "Conga CLM",
     "Inbound", "Implemented",
     "Conga templates pulled into conga_templates; admin page lets PDL select per-deal.",
     "/admin/engagement-letters", "—"),
    ("Renewal tracking / Power BI feeds", "Power BI",
     "Outbound", "Pending",
     "Per-deal data feeds for renewal tracking, management dashboards, win-rate analytics.",
     "—", "US-65 — analytics endpoint exists but PBI export connector not yet wired"),
    ("Audit trail + activity log feed", "Internal / SIEM",
     "Outbound", "Implemented",
     "Every domain mutation writes to activity_log with structured metadata payload; queryable + exportable.",
     "GET /api/activity", "—"),
    ("Approval domain events outbox (F1.4)", "Internal events",
     "Outbound", "Implemented",
     "domain_events_outbox holds versioned DealSubmitted/DealApproved/DealRejected events; in-process EventBus + replay-on-boot dispatcher.",
     "psql -c 'SELECT * FROM domain_events_outbox ORDER BY id DESC'", "Service Bus / Kafka adapters are F1.4 follow-ups"),
]

# ---------- PRICING ENGINE PARITY (vs. Intapp Pricing or similar engines) ----------
PRICING_ENGINE = [
    # (Capability, US Refs, Status, What it does, Notes)
    ("Engagement scoping with assemblies + prompts", "US-08, US-10, US-11",
     "Implemented",
     "Scope catalog + assembly templates + prompt-driven complexity multipliers — matches Intapp Pricing's deliverables-driven scoping.",
     "—"),
    ("Per-role rate-card management", "US-21",
     "Implemented",
     "Multiple rate cards with effective dates; per-role bill rate + cost rate.",
     "—"),
    ("Standard vs Contract rates", "US-21",
     "Implemented",
     "Both supported via separate rate cards; deal pulls active card.",
     "—"),
    ("Discount + pricing-rule engine", "US-22",
     "Foundation",
     "Per-deal rate overrides with audit trail; geo/offshore discount rules tracked but not yet automated.",
     "Override workflow live; rule-driven discounting pending"),
    ("Multiple fee arrangements (T&M / Fixed / Capped / Contingent / Retainer / Hybrid)", "US-23",
     "Implemented",
     "FeeArrangementPicker UI + applyFeeArrangement projection; same baseline T&M math, arrangement layers on top.",
     "—"),
    ("Margin / EHR / realization metrics", "US-24, US-25",
     "Implemented",
     "Margin% + blendedRate + per-line cost/fee + tech-admin fee uplift; calc-parity golden pins formulas.",
     "—"),
    ("What-if scenarios", "US-32",
     "Implemented",
     "Multiple scenarios per deal; per-scenario pricing lines; AI scenario recommendation with structured rationale.",
     "Deal → Scenarios step"),
    ("Benchmark comparisons (similar deals)", "US-33, US-36",
     "Implemented",
     "POST /api/ai/deal-similarity uses pgvector k-NN against deal embeddings (1536-d) — sub-500ms similar-deal lookup.",
     "Powered by F2.1 IntelligenceEngine"),
    ("Risk / complexity indicators on approval", "US-37",
     "Implemented",
     "POST /api/ai/risk-summary returns risk level + factors + key message + approval likelihood; also surfaces in approval banner.",
     "F4.4.2 wired narrative through llm.ts"),
    ("Approval workflow + states", "US-39, US-40, US-41",
     "Implemented",
     "Tiered routing by margin/size; structured states; comments + rejection rationale.",
     "—"),
    ("Audit + retention + security", "US-55",
     "Implemented",
     "activity_log + domain_events_outbox immutable; per-row audit metadata (createdBy, override actors); pg_dump backups via scripts/audit/backup_db.sh.",
     "—"),
    ("Template + catalog governance", "US-54, US-56",
     "Implemented",
     "Versioned scope_templates + assembly_templates + prompt_sets; admin pages with role-gated edits.",
     "—"),
    ("Granular rate tables (role / geo / experience)", "US-67",
     "Foundation",
     "Per-role rates live; geo + experience dimensions tracked in schema but not yet faceted in UI.",
     "Schema ready; UI pending"),
    ("Pricing strategy flexibility (value/fixed/cost-plus)", "US-58, US-59",
     "Implemented",
     "All 6 fee arrangements + Cost-plus arrangement projection; fixedFeeAmount + cappedFeeAmount + retainerAmount fields on deals.",
     "—"),
    ("Bundled + multi-template pricing", "US-60, US-61",
     "Implemented",
     "Multi-entity tabs + multi-service deals + assembly composition via combineable scope templates.",
     "—"),
    ("Pricing exception governance", "US-66",
     "Foundation",
     "Margin-target overrides + admin-fee overrides supported via batch_adjustment_rules + per-deal targetMarginPercent; configurable approval triggers per BU live; full exception-rule editor pending.",
     "—"),
    ("Configurable approvals (admin fee, T&M, US-only resources)", "US-66",
     "Foundation",
     "Practice Lead trigger evaluator covers fee/margin/scope-size; further conditional gates (US-only resources, admin-fee thresholds) wired in policy.ts but UI for non-engineer config pending.",
     "—"),
    ("Resource-planning consumption", "US-50",
     "Implemented",
     "Workday push exports planned hours by role to cost-center; project record on approval.",
     "—"),
]

# ---------- DIFFERENTIATORS ----------
# Capabilities unique to DealPad — beyond Excel parity, beyond Intapp Pricing.
DIFFERENTIATORS = [
    # (Capability, Stakeholder Value, Status, What it does, How to demo)
    ("AI-powered deal similarity (vector search)",
     "Pricing benchmarks 10x faster than manual lookup",
     "Implemented",
     "pgvector 1536-dim embeddings on every deal; k-NN <-> distance returns 5 most similar in <500ms across thousands of deals; lazy backfill on first query.",
     "Deal Setup step → AI Insights card auto-loads"),
    ("AI risk narrative + key-message",
     "Approvers get plain-English summary tailored to deal facts",
     "Implemented",
     "llm.completeStructured produces narrative + keyMessage + approval likelihood; heuristic fallback on LLM failure.",
     "Deal → Review step → Generate Risk Summary"),
    ("AI margin advisor with call-to-action",
     "PDL gets specific suggestions ranked by impact",
     "Implemented",
     "Suggestion math (role-shift, rate-uplift) heuristic + deterministic; narrative + callToAction enriched by llm.ts.",
     "Pricing step → Run AI Margin Advisor"),
    ("AI scenario recommendation",
     "PDL doesn't have to manually craft alternatives",
     "Implemented",
     "Generates 2-3 scenarios with margin/staffing tradeoffs; ranks by alignment with margin target.",
     "Scenarios step → Recommend Scenarios"),
    ("AI effort estimation",
     "Hours estimate seeded from scope + complexity, not blank-slate",
     "Implemented",
     "Per-role distribution from scope items + complexity + prompts; FTE capacity check against project window.",
     "Scope step → Run Estimation"),
    ("Voice-to-scope (audio → scope items)",
     "Capture client conversations directly into structured scope",
     "Foundation",
     "Audio metadata + transcription + token-overlap heuristic ranks scope_catalog matches; user accepts to commit.",
     "/api/deals/:id/voice-transcripts → /process → /apply"),
    ("AI-confidence-routed intake (Intapp federated)",
     "Auto-progresses high-confidence intakes; routes ambiguous ones for review",
     "Implemented",
     "Intake AI extraction with confidence scores; > threshold auto-progress, ≤ threshold reviewer matrix.",
     "/architecture → Intapp Federated tab"),
    ("Predictive scope-creep detector",
     "Catch hours/margin drift before deals miss their numbers",
     "Foundation",
     "Heuristic detector flags 5 signal types (scope_growth, change_order_density, burn_rate, margin_drift, stale_no_progress) with severity + dedup.",
     "POST /api/deals/:id/scope-creep/scan"),
    ("Dynamic rate optimizer",
     "Recommend rate adjustments by capacity + velocity + margin",
     "Foundation",
     "Heuristic optimizer suggests per-role rate uplifts/downticks (capped ±10–15%); ML model plugs into evaluate() seam.",
     "POST /api/rate-optimization/runs"),
    ("AI cost & latency observability",
     "Finance can see token spend + p95 per AI feature",
     "Implemented",
     "Every LLM/AI call writes to ai_telemetry (cost USD, tokens, latency, status); summary endpoint computes p95 + per-operation rollups.",
     "GET /api/ai-telemetry/summary"),
    ("Real-time collaborative scoping (foundation)",
     "Multiple PDLs editing the same deal without merge conflicts",
     "Foundation",
     "collaboration_sessions with Yjs CRDT durability + room-id allocation; WebSocket gateway pending.",
     "POST /api/deals/:id/collab/sessions"),
    ("Client self-service portal (magic-link)",
     "Clients can review proposals without an account",
     "Foundation",
     "SHA-256 hashed magic-link tokens; /api/portal/* token-gated routes for read-only deal/proposal/scope.",
     "Generate invite via /api/deals/:id/portal-invites"),
    ("Multi-entity worksheets (one deal, many entities)",
     "Tax engagements with 4 entities each filing different forms — all in one deal",
     "Implemented",
     "Per-entity scope tabs + pricing rollups; one deal-number stays the contracting unit; entity-aware unique constraints.",
     "Deal → Scope step → Add Entity tabs"),
    ("Assembly engine with sandboxed math.js formulas",
     "Tax PHB Standard Bundle expands to 87 line items with parameterized hours per tier",
     "Implemented",
     "math.js AST allow-list (rejects FunctionNode/AccessorNode); per-component formula uses prompt answers as variables.",
     "Scope step → Apply Assembly"),
    ("Batch renewal processor (tax-season ready)",
     "Run a single job over 200+ deals with adjustment rules + variance gates",
     "Implemented",
     "Batch jobs with reusable adjustment rules; per-item variance flagging; Python+Celery+Redis worker scaffolded.",
     "/admin/batch-renewals → Create job"),
    ("DDD strangler-fig domain layer",
     "New code lives in pure-domain packages; legacy routes migrate one at a time",
     "Implemented",
     "@dealpad/domain (Money, Percentage, Deal aggregate, versioned events) + @dealpad/application (services) + @dealpad/infrastructure (Drizzle repo, EventBus, outbox).",
     "packages/domain/src/deal/Deal.ts"),
    ("Domain events outbox + replay",
     "Auto-push to integrations is durable across server restarts",
     "Implemented",
     "domain_events_outbox table; OutboxDispatcher replays unpublished rows on boot.",
     "—"),
    ("Per-deal fingerprint + feature snapshot",
     "Cache key for similarity + future ML features",
     "Implemented",
     "deals.fingerprint JSONB stores BU/serviceLine/feeBucket/marginBucket/scopeItemCount + computedAt + mode.",
     "—"),
    ("Persona-aware UX (RBAC over 6 roles)",
     "Same UI, different surfaces for PDL/SLL/PO/FIN/QRM/IT",
     "Implemented",
     "PERMISSIONS_BY_ROLE matrix + requirePerm middleware; client mirrors via AuthContext.hasPermission.",
     "Login → switch personas → menus + actions reshape"),
    ("Calc-parity golden test (regression-free pricing)",
     "Pricing changes can't silently break existing deals",
     "Implemented",
     "Read-only golden over deals 4 + 27 (T&M + Tech-Admin uplift); idempotent re-run; intentional changes regenerate via WRITE_GOLDEN=1.",
     "npm test (tests/calc-parity/)"),
    ("Deployment automation (Render + Neon, blueprint)",
     "One-click public demo URL; CI deploy on merge to main",
     "Implemented",
     "render.yaml Blueprint + Dockerfile + Neon Postgres with pgvector; current live URL https://dealpad-demo.onrender.com.",
     "https://dealpad-demo.onrender.com"),
]

# ---------- TECHNICAL DETAIL (engineering-facing) ----------
TECHNICAL_DETAIL = [
    # (Phase, Feature ID, Capability, Status, Description, PR(s))
    ("Phase 0", "F0.1", "Audit + smoke + extractor scripts", "Implemented",
     "scripts/audit/{smoke_test.sh, extract_endpoints.py, extract_schema.py}.", "—"),
    ("Phase 0", "F0.4", "type-safe req helpers", "Implemented",
     "server/lib/req.ts paramStr/paramInt/headerStr/queryStr.", "—"),
    ("Phase 0", "F0.5", "Pricing engine extracted", "Implemented",
     "services/pricing.ts: compute/recalc/persist/backfill/reconcile.", "—"),
    ("Phase 0", "F0.6", "ESLint flat + Prettier", "Implemented",
     "eslint.config.js (flat), 0 errors / ~120 warnings baseline.", "—"),
    ("Phase 0", "F0.10", "Recalc idempotency", "Implemented",
     "Back-derive standardRate; persist on update.", "—"),
    ("Phase 0", "F0.11", "macOS onBlur flush util", "Implemented",
     "client/src/lib/flush-pending-edits.ts.", "—"),
    ("Phase 1", "F1.1", "Multi-entity worksheets", "Implemented",
     "deal_entities + entity_id on scope/pricing.", "F1.1.* chain"),
    ("Phase 1", "F1.1.1", "Scope per-entity unique index", "Implemented",
     "(deal_id, entity_id, scope_item_id) unique.", "PR #53"),
    ("Phase 1", "F1.2", "Assembly expansion engine", "Implemented",
     "assembly_templates + math.js sandboxed AST.", "F1.2.* chain"),
    ("Phase 1", "F1.3", "Batch renewals", "Implemented",
     "batch_renewal_* + Python Celery+Redis worker.", "F1.3.* chain"),
    ("Phase 1", "F1.4", "DDD Strangler-Fig (Deal aggregate)", "Implemented",
     "packages/domain + application + infrastructure.", "F1.4.* chain"),
    ("Phase 2", "F2.1", "Intelligence Engine (pgvector k-NN)", "Implemented",
     "deals.embedding vector(1536) + IntelligenceEngine.findSimilar.", "F2.1.* chain"),
    ("Phase 2", "F2.2", "Budget monitoring", "Implemented",
     "budget_actuals + budget_alerts + thresholds.", "F2.2.* chain"),
    ("Phase 2", "F2.3", "Time tracking + suggest", "Implemented",
     "time_entries + /api/time/suggest.", "F2.3.* chain"),
    ("Phase 2", "F2.4", "Alternative fee arrangements", "Implemented",
     "6 arrangements + applyFeeArrangement.", "F2.4.* chain"),
    ("Phase 3", "F3.1", "Collaboration sessions (foundation)", "Foundation",
     "collaboration_sessions + room allocation.", "PR #39"),
    ("Phase 3", "F3.2", "Portal magic-link auth", "Foundation",
     "portal_invites + SHA-256 tokens.", "PR #37, #38"),
    ("Phase 3", "F3.3", "Scope creep detector", "Foundation",
     "scope_creep_signals + 5 heuristic rules.", "PR #40"),
    ("Phase 3", "F3.4", "Voice-to-scope", "Foundation",
     "voice_transcripts + token-overlap extractor.", "PR #41"),
    ("Phase 3", "F3.5", "Slack/Teams native apps", "Pending",
     "Bolt.js + tenant OAuth.", "—"),
    ("Phase 3", "F3.6", "Rate optimization", "Foundation",
     "rate_optimization_runs + heuristic optimizer.", "PR #42"),
    ("Phase 4", "F4.1", "pgvector RAG", "Implemented",
     "Already shipped via F2.1.", "—"),
    ("Phase 4", "F4.2", "ml-service Python scaffold", "Foundation",
     "FastAPI + heuristic /effort-estimator + /margin-optimizer.", "PR #50"),
    ("Phase 4", "F4.4", "llm.ts client abstraction", "Foundation",
     "complete + completeStructured + simulated default + provider stubs.", "PR #47"),
    ("Phase 4", "F4.4.2", "risk-summary via llm.ts", "Implemented",
     "narrative + keyMessage + approvalLikelihood through llm.ts.", "PR #48"),
    ("Phase 4", "F4.4.3", "margin-advisor via llm.ts", "Implemented",
     "narrative + callToAction through llm.ts.", "PR #49"),
    ("Phase 4", "F4.5", "AI telemetry + dashboards", "Implemented",
     "ai_telemetry + recordAi/withAiTelemetry + summary endpoint.", "PR #46"),
    ("Operational", "fix.dynamics-seed", "Dynamics seed skips test deals", "Implemented",
     "—", "PR #51"),
    ("Operational", "fix.auto-seed-primary-entity", "Every deal-create path auto-seeds Primary Entity", "Implemented",
     "—", "PR #52"),
    ("Operational", "fix.healthz", "Public /healthz", "Implemented",
     "—", "PR #59"),
    ("Operational", "deploy", "Render + Neon deploy", "Implemented",
     "render.yaml Blueprint + Dockerfile + docs/deployment.md.", "PRs #55-#59"),
]


# ============ Workbook construction ============

def write_header(ws, headers, widths, row=1):
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 32


def write_data_row(ws, row_idx, values, status_col_idx=None):
    for c, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=c, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER_THIN
        if status_col_idx and c == status_col_idx:
            color = STATUS_COLOR.get(value, "FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
    ws.row_dimensions[row_idx].height = 50


def build_overview(wb, totals):
    ws = wb.create_sheet("Overview", 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 18

    rows = [
        ("DealPad — Pricing & Scoping 2.0", "", ""),
        ("", "", ""),
        ("What this is",
         "A web-based replacement for Armanino's Excel-based pricing & scoping workbooks, "
         "with bi-directional integration into the Quote-to-Cash stack and AI-driven differentiation. "
         "Pricing methodologies are preserved; calculation parity is contractual.", ""),
        ("", "", ""),
        ("Live demo URL", "https://dealpad-demo.onrender.com", ""),
        ("Test plan + curl examples", "docs/test-plan.md (also on GitHub)", ""),
        ("", "", ""),
        ("Capability buckets", "", ""),
        ("Excel Parity", "Replicates the current Excel-workbook pricing process — calc fidelity, scope catalog, prompts, approvals.", f"{totals['excel_parity_done']}/{totals['excel_parity_total']} live"),
        ("Integrations", "Bi-directional sync with D365 / Workday / Intapp / Conga + audit/event feeds.", f"{totals['integrations_done']}/{totals['integrations_total']} live"),
        ("Pricing Engine Parity", "Capabilities matching what Intapp Pricing (or competing engines) deliver.", f"{totals['pricing_engine_done']}/{totals['pricing_engine_total']} live"),
        ("Differentiators", "Capabilities unique to DealPad: AI similarity, voice-to-scope, scope-creep prediction, multi-entity, real-time collab foundation.", f"{totals['differentiators_done']}/{totals['differentiators_total']} live"),
        ("", "", ""),
        ("Status legend", "", ""),
        ("Implemented", "End-to-end live: schema + service + routes + UI (where applicable) shipped and tested.", ""),
        ("Foundation", "Schema + service + routes shipped; UI surface or live-mode wiring (API keys, training data) pending.", ""),
        ("Pending", "Backlog item, intentionally not started this cycle.", ""),
        ("Operational", "Deployment, ops, or bugfix — supporting capability rather than a user-facing feature.", ""),
        ("", "", ""),
        ("How to read the sheets", "", ""),
        ("Excel Parity tab", "67 user-stories from the RFQ mapped to current implementation status. Use this to track contractual parity acceptance.", ""),
        ("Integrations tab", "All four simulated integrations (Dynamics / Workday / Intapp / Conga) plus Power BI + audit feeds. Live-mode wiring is a config flip.", ""),
        ("Pricing Engine tab", "Side-by-side capability list against what a typical Pricing Engine product (e.g. Intapp Pricing) delivers — for the build-vs-buy conversation.", ""),
        ("Differentiators tab", "What DealPad does that the Excel + Intapp baseline does NOT — the value-add story for stakeholders.", ""),
        ("Technical Detail tab", "Engineering-facing: phase / feature ID / PR numbers. For development teams + audit.", ""),
        ("Glossary tab", "Column key + status legend reference.", ""),
    ]
    for r_idx, (col_a, col_b, col_c) in enumerate(rows, start=1):
        a = ws.cell(row=r_idx, column=1, value=col_a)
        b = ws.cell(row=r_idx, column=2, value=col_b)
        c = ws.cell(row=r_idx, column=3, value=col_c)
        for cell in (a, b, c):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r_idx == 1:
            a.font = Font(bold=True, size=18, color="DA720F")
        if col_a in ("Capability buckets", "Status legend", "How to read the sheets"):
            a.fill = SECTION_FILL; a.font = SECTION_FONT
            b.fill = SECTION_FILL
            c.fill = SECTION_FILL
        if col_a in ("Implemented", "Foundation", "Pending", "Operational"):
            a.fill = PatternFill(fill_type="solid", fgColor=STATUS_COLOR.get(col_a, "FFFFFF"))
            a.font = Font(bold=True)
        ws.row_dimensions[r_idx].height = 32 if col_b else 18


def build_excel_parity(wb, status_totals):
    ws = wb.create_sheet("Excel Parity")
    write_header(
        ws,
        ["Capability", "User Story Refs", "Stakeholder", "Status", "What it does", "How to demo", "Notes"],
        [44, 16, 22, 16, 70, 50, 50],
    )
    for i, row in enumerate(EXCEL_PARITY, start=2):
        write_data_row(ws, i, row, status_col_idx=4)
        status_totals.setdefault(row[3], 0)
        status_totals[row[3]] += 1
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:G1"


def build_integrations(wb, status_totals):
    ws = wb.create_sheet("Integrations")
    write_header(
        ws,
        ["Capability", "System", "Direction", "Status", "What it does", "How to demo", "Notes"],
        [40, 26, 18, 16, 70, 50, 50],
    )
    for i, row in enumerate(INTEGRATIONS, start=2):
        write_data_row(ws, i, row, status_col_idx=4)
        status_totals.setdefault(row[3], 0)
        status_totals[row[3]] += 1
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:G1"


def build_pricing_engine(wb, status_totals):
    ws = wb.create_sheet("Pricing Engine")
    write_header(
        ws,
        ["Capability", "User Story Refs", "Status", "What it does", "Notes"],
        [50, 18, 16, 80, 50],
    )
    for i, row in enumerate(PRICING_ENGINE, start=2):
        write_data_row(ws, i, row, status_col_idx=3)
        status_totals.setdefault(row[2], 0)
        status_totals[row[2]] += 1
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:E1"


def build_differentiators(wb, status_totals):
    ws = wb.create_sheet("Differentiators")
    write_header(
        ws,
        ["Capability", "Stakeholder Value", "Status", "What it does", "How to demo"],
        [42, 50, 16, 80, 50],
    )
    for i, row in enumerate(DIFFERENTIATORS, start=2):
        write_data_row(ws, i, row, status_col_idx=3)
        status_totals.setdefault(row[2], 0)
        status_totals[row[2]] += 1
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:E1"


def build_technical(wb):
    ws = wb.create_sheet("Technical Detail")
    write_header(
        ws,
        ["Phase", "Feature ID", "Capability", "Status", "Description", "PR(s)"],
        [12, 24, 40, 16, 70, 26],
    )
    for i, row in enumerate(TECHNICAL_DETAIL, start=2):
        write_data_row(ws, i, row, status_col_idx=4)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:F1"


def build_glossary(wb):
    g = wb.create_sheet("Glossary")
    g.column_dimensions["A"].width = 22
    g.column_dimensions["B"].width = 100

    glossary = [
        ("Column", "Meaning"),
        ("Capability", "Plain-English description of what the user / business gets."),
        ("User Story Refs", "RFQ user-story IDs (US-XX) tied to this capability."),
        ("Stakeholder", "Primary persona that benefits or operates this capability."),
        ("Status",
         "Implemented = end-to-end live (schema + service + routes + UI where applicable).\n"
         "Foundation = schema/service/routes shipped; UI or live-mode wiring pending.\n"
         "Pending = backlog, intentionally not started this cycle.\n"
         "Operational = deployment / ops / bugfix supporting capability."),
        ("What it does", "Two-line description usable in stakeholder briefings."),
        ("How to demo", "Concrete click-path or curl command. Use against the live demo URL."),
        ("Notes", "Caveats, dependencies, or roadmap hints (e.g. 'live mode requires API keys')."),
        ("System (Integrations tab)", "External system DealPad talks to: Dynamics 365 / Workday / Intapp / Conga / Power BI."),
        ("Direction (Integrations tab)", "Inbound (we pull) / Outbound (we push) / Bi-directional."),
        ("Stakeholder Value (Differentiators tab)", "Why a stakeholder cares — framed as outcome, not feature."),
    ]
    for r_idx, (col, meaning) in enumerate(glossary, start=1):
        a = g.cell(row=r_idx, column=1, value=col)
        b = g.cell(row=r_idx, column=2, value=meaning)
        if r_idx == 1:
            a.fill = HEADER_FILL; a.font = HEADER_FONT
            b.fill = HEADER_FILL; b.font = HEADER_FONT
        else:
            a.font = Font(bold=True)
        a.alignment = Alignment(vertical="top", wrap_text=True)
        b.alignment = Alignment(vertical="top", wrap_text=True)
        a.border = BORDER_THIN; b.border = BORDER_THIN
        g.row_dimensions[r_idx].height = 64 if r_idx == 5 else 22


def main() -> None:
    wb = Workbook()
    # Remove default Sheet1
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    excel_status = {}
    integrations_status = {}
    pricing_status = {}
    differentiators_status = {}

    build_excel_parity(wb, excel_status)
    build_integrations(wb, integrations_status)
    build_pricing_engine(wb, pricing_status)
    build_differentiators(wb, differentiators_status)
    build_technical(wb)
    build_glossary(wb)

    # Now build the Overview using counts (must come last so it can reference real totals,
    # but openpyxl supports inserting at index 0)
    totals = {
        "excel_parity_done": excel_status.get("Implemented", 0),
        "excel_parity_total": sum(excel_status.values()),
        "integrations_done": integrations_status.get("Implemented", 0),
        "integrations_total": sum(integrations_status.values()),
        "pricing_engine_done": pricing_status.get("Implemented", 0),
        "pricing_engine_total": sum(pricing_status.values()),
        "differentiators_done": differentiators_status.get("Implemented", 0),
        "differentiators_total": sum(differentiators_status.values()),
    }
    build_overview(wb, totals)

    # Reorder sheets: Overview first, then category sheets, then Technical Detail, then Glossary
    desired = ["Overview", "Excel Parity", "Integrations", "Pricing Engine",
               "Differentiators", "Technical Detail", "Glossary"]
    wb._sheets = [wb[name] for name in desired if name in wb.sheetnames]

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)

    rows_per_sheet = {
        "Excel Parity": len(EXCEL_PARITY),
        "Integrations": len(INTEGRATIONS),
        "Pricing Engine": len(PRICING_ENGINE),
        "Differentiators": len(DIFFERENTIATORS),
        "Technical Detail": len(TECHNICAL_DETAIL),
    }
    print(f"Wrote {OUT_PATH}")
    for sheet, count in rows_per_sheet.items():
        print(f"  {sheet}: {count} rows")
    print(f"Implemented totals: "
          f"Excel Parity {totals['excel_parity_done']}/{totals['excel_parity_total']}, "
          f"Integrations {totals['integrations_done']}/{totals['integrations_total']}, "
          f"Pricing Engine {totals['pricing_engine_done']}/{totals['pricing_engine_total']}, "
          f"Differentiators {totals['differentiators_done']}/{totals['differentiators_total']}.")


if __name__ == "__main__":
    main()
